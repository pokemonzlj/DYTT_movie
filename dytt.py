import requests
from bs4 import BeautifulSoup as bsp
import re
import openpyxl
from openpyxl import load_workbook
import os

site = 'https://www.dytt8.net'

class DYTT_Movie:

    def __init__(self):
        pass


def getSoup(url):
    response = requests.get(url)
    response.encoding = 'gb18030'  # 只能用这个编码，否则中文乱码
    # response.encoding = 'utf-8'
    return bsp(response.text, "html.parser")


def get_movie_detail(url):
    soup = getSoup(url)
    # 先将电影内容信息整个提取出来
    # 查找<td>标签并获取其文本内容
    td_text = soup.find_all('td')
    cleaned_text = ''
    for td in td_text:
        text = td.get_text(strip=True)  # strip=True会去除多余的空白字符
        text = text.replace('　', '')
        if text:
            # 去除所有空格（包括换行符、制表符等）并转换为大写
            cleaned_text += ''.join(text.split()).upper()
    # 提取电影信息
    # print(cleaned_text)
    # 提取年代
    year_match = re.search(r'◎年代(\d+)', cleaned_text)
    year = year_match.group(1) if year_match else 'N/A'
    # 提取产地
    country_match = re.search(r'◎产地(.*?)◎', cleaned_text)
    country = country_match.group(1).strip() if country_match else 'N/A'
    # 提取类别
    genre_match = re.search(r'◎类别(.*?)◎', cleaned_text)
    genre = genre_match.group(1).strip() if genre_match else 'N/A'
    genre = genre.replace('/', '')
    # 提取语言
    language_match = re.search(r'◎语言(.*?)◎', cleaned_text, re.DOTALL)
    language = language_match.group(1).strip() if language_match else 'N/A'
    language = language.replace('/', '')
    # 提取字幕
    subtitle_match = re.search(r'◎字幕(.*?)◎', cleaned_text, re.DOTALL)
    subtitle = subtitle_match.group(1).strip() if subtitle_match else 'N/A'
    # 提取IMDb评分
    imdb_score_match = re.search(r'IMDB评分(.*?)FROM', cleaned_text, re.DOTALL)
    imdb_rating = imdb_score_match.group(1).split('/')[0].strip() if imdb_score_match else 'N/A'
    # 提取豆瓣评分
    douban_score_match = re.search(r'豆瓣评分(.*?)FROM', cleaned_text, re.DOTALL)
    douban_rating = douban_score_match.group(1).split('/')[0].strip() if douban_score_match else 'N/A'

    # 打印电影信息
    print(f"年代：{year}")
    print(f"产地：{country}")
    print(f"类别：{genre}")
    print(f"语言：{language}")
    print(f"字幕：{subtitle}")
    print(f"IMDb评分：{imdb_rating}")
    print(f"豆瓣评分：{douban_rating}")
    if douban_rating != 'N/A':
        rating = douban_rating
    elif imdb_rating != 'N/A':
        rating = imdb_rating
    else:
        rating = '0'
    if subtitle == 'N/A':  # 没返回字幕的通常就是中文字幕
        subtitle = '中文字幕'
    elif subtitle == '中文':
        subtitle = "中文字幕"
    elif subtitle == '中英双字':
        subtitle = "中英双字幕"
    return year, country, genre, language, subtitle, rating
    # year = soup.find(string=lambda text: '年' in text and '2024' in text).strip().split('年')[0].strip()
    # country = soup.find(string='◎产地').find_next_sibling(string=True).strip()
    # genre = soup.find(string='◎类别').find_next_sibling(string=True).strip()
    # language = soup.find(string='◎语言').find_next_sibling(string=True).strip()
    # subtitle = soup.find(string='◎字幕').find_next_sibling(string=True).strip()
    # imdb_score = soup.find(string=lambda text: 'IMDB评分' in text)
    # imdb_rating = imdb_score.find_next_sibling(string=True).split('from')[0].strip() if imdb_score else 'N/A'
    # douban_score = soup.find(string=lambda text: '豆瓣评分' in text)
    # douban_rating = douban_score.find_next_sibling(string=True).split('from')[0].strip() if douban_score else 'N/A'
    #
    # # 打印电影信息
    # print(f"年代：{year}")
    # print(f"产地：{country}")
    # print(f"类别：{genre}")
    # print(f"语言：{language}")
    # print(f"字幕：{subtitle}")
    # print(f"IMDb评分：{imdb_rating}")
    # print(f"豆瓣评分：{douban_rating}")


def deal_movie_name(str):
    start_index = str.find("《")
    if start_index != -1:  # 确保找到了《
        # 找到》的索引
        end_index = str.find("》", start_index)
        if end_index != -1:  # 确保找到了》
            # 提取尖括号之间的内容
            movie_name = str[start_index + 1:end_index]
            return movie_name
    return str


def get_movie(url):
    soup = getSoup(url)
    movie_tables = soup.find_all('table', class_='tbspan')  # 查找所有的电影内容结构体
    # 遍历电影表格
    movie_list = []
    for table in movie_tables:
        # 查找电影名称的链接
        movie_link = table.find('a', class_='ulink')
        if movie_link:
            movie_info = {}
            movie_name = movie_link.get_text()  # 获取电影名称
            movie_url = movie_link['href']  # 获取电影详情地址
            detail_url = site + movie_url
            print(f"电影名称: {movie_name}")
            movie_name = deal_movie_name(movie_name)
            movie_info['电影名称'] = movie_name
            print(f"电影详情地址: {detail_url}")
            movie_info['电影详情地址'] = detail_url
            year, country, genre, language, subtitle, rating = get_movie_detail(detail_url)
            movie_info['年代'] = year
            movie_info['产地'] = country
            movie_info['类别'] = genre
            movie_info['语言'] = language
            movie_info['字幕'] = subtitle
            movie_info['评分'] = rating
            if "/" in movie_name:
                parts = movie_name.split('/')
                movie_name = parts[0]

            splice_movie_name = movie_name + year + "年" + genre + "片评分" + rating + language + subtitle
            movie_info['本地视频文件名'] = splice_movie_name
            print(f"本地视频文件名: {splice_movie_name}")
            print("-" * 80)
            movie_list.append(movie_info)
    return movie_list


def initialization_excel():
    if not os.path.exists('电影清单.xlsx'):
        wb = openpyxl.Workbook()
        sheetname = wb.sheetnames
        sheet = wb[sheetname[0]]
        title_list = ['电影名称', '电影详情地址', '年代', '产地', '类别', '语言', '字幕', '评分', '本地视频文件名']
        for i in range(1, len(title_list) + 1):
            sheet.cell(row=1, column=i).value = title_list[i - 1]
        wb.save('电影清单.xlsx')


def write_movie_to_excel(movie_list=[]):
    initialization_excel()
    wb = load_workbook('电影清单.xlsx')
    sheet = wb.active  # 获取工作表
    max_row = sheet.max_row  # 获取最大行数
    movies_already_have = []
    for row_offset,row in enumerate(sheet.iter_rows(min_col=1, max_col=1, values_only=True)):
        if row_offset > 1:
            movies_already_have.append(row[0])
    for movie in movie_list:
        if movie['电影名称'] not in movies_already_have:
            sheet.cell(row=max_row + 1, column=1).value = movie['电影名称']
            sheet.cell(row=max_row + 1, column=2).value = movie['电影详情地址']
            sheet.cell(row=max_row + 1, column=3).value = movie['年代']
            sheet.cell(row=max_row + 1, column=4).value = movie['产地']
            sheet.cell(row=max_row + 1, column=5).value = movie['类别']
            sheet.cell(row=max_row + 1, column=6).value = movie['语言']
            sheet.cell(row=max_row + 1, column=7).value = movie['字幕']
            sheet.cell(row=max_row + 1, column=8).value = movie['评分']
            sheet.cell(row=max_row + 1, column=9).value = movie['本地视频文件名']
            max_row += 1
    wb.save('电影清单.xlsx')


def total_loop(num=5):
    for index in range(num):
        index += 1
        url = 'http://www.dytt8.net/html/gndy/dyzz/list_23_' + \
              str(index) + '.html'
        movie_list = get_movie(url)
        write_movie_to_excel(movie_list)


if __name__ == '__main__':
    total_loop(2)
